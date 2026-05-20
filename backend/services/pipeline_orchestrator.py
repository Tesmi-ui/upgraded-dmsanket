"""
backend/intelligence_engine.py
================================
BRLF v4.1 — Full Pipeline Orchestrator.

Two public APIs:
  run_pipeline(ProcessRequest) → ProcessResponse
      Field-level validation + inference, used by POST /api/process.

  DataMigrationEngine(config).process(file_path, out_dir) → dict
      Full file migration pipeline: dedup, spell-check, auto-correct,
      9-sheet audit report. Used by BulkProcessor and /api/upload jobs.

FIXES vs v4.0
-------------
  FIX 1:  infer_gender/infer_category imported from gender_lib and
          check_intelligence.inference.engine (single source of truth).
  FIX 2:  FEMALE_NAMES etc. imported from gender_lib.knowledge_base.
  FIX 3:  NameSpellChecker in validators/name_spell_checker.py.
  FIX 4:  _auto_correct() NaN check replaced with _clean_val() helper.
  FIX 5:  _remove_duplicates() guards against _sort_date column collision.
  FIX 6:  _validate_output() expected count accounts for n_placeholder.
  FIX 7:  _build_report() Sheet 9 guards against empty spell_df (no columns).
  FIX 8:  process() step numbering corrected (two steps were both '# 9.').
  FIX 9:  check_series() token_results reused from first call (no 2x processing).
  FIX 10: All imports point to single sources — no duplicate knowledge bases.
  FIX 11: Returns raw dict from DataMigrationEngine; run_pipeline returns
          ProcessResponse — callers choose what they need.

  BUG FIX (v4.1-a): _auto_correct() pandas index crash — enumerate(results)
          replaced with results.items() in gender and category blocks.
          reset_index(drop=True) added at top of _auto_correct() as belt-and-
          suspenders guard. Fixes ValueError: 30682 is not in range on merged
          multi-file datasets where concat() leaves non-contiguous index labels.

  BUG FIX (v4.1-b): _build_review_file() final log.info used bare name
          'merged_indices' (NameError). Replaced with 'merged_positions'
          which is the set defined 30 lines above in the same method.

  BUG FIX (v4.1-c): _infer_category() now passes district column value to
          engine.infer_category() so Tier-4 geographic prior activates for
          Gadchiroli, Nandurbar etc. Added self.district_col config key.

  CATEGORY UPGRADE (v4.1): infer_category now uses 4-tier pipeline from
          engine.py v4.1 — T1 exact (conf 88/62), T2 fuzzy (74), T3 all-token
          (68), T4 geographic (55). AMBIGUOUS_SURNAMES capped at conf 62.
"""

from __future__ import annotations

import logging
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from core.execution_context import ExecutionContext, ExecutionMode

import pandas as pd

from gender_lib import infer_gender as _lib_infer_gender
from gender_lib.knowledge_base import (
    FEMALE_NAMES, MALE_NAMES,
    FEMALE_SUFFIXES, MALE_SUFFIXES,
    SURNAME_CATEGORY,
)
from check_intelligence.inference.engine import (
    infer_gender   as _brlf_infer_gender,
    infer_category as _brlf_infer_category,
)
from validators.name_formatter     import NameCaseFormatter
from validators.contact_validator  import ContactValidator
from validators.mgnrega_validator  import MGNREGAValidator
from validators.name_spell_checker import NameSpellChecker
from models.schema import (
    FarmerRecord, ProcessedRecord, ProcessRequest, ProcessResponse,
    ValidationDetail, InferenceDetail,
)

log = logging.getLogger("brlf.intelligence_engine")

CAT_MIN_CONFIDENCE = 60

PLACEHOLDER_IDS = {"v1", "version 1", "version1", "test", "null", "none", "na", "n/a", ""}

_name_fmt  = NameCaseFormatter()
_contact_v = ContactValidator()
_mgnrega_v = MGNREGAValidator()


# ─────────────────────────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────────────────────────

def _clean_val(v) -> str:
    """Return stripped lowercase str, or '' for None / NaN / empty."""
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return str(v).strip().lower()


# ─────────────────────────────────────────────────────────────────────────────
# RECORD-LEVEL PIPELINE  (/api/process)
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline(
    request: ProcessRequest,
    context: Optional[ExecutionContext] = None,
) -> ProcessResponse:
    """
    Process a batch of FarmerRecord objects.
    Returns ProcessResponse with per-row results and summary counts.

    context controls which rules / columns are allowed to mutate values.
    If context is None → GLOBAL mode (backward-compatible with v4.0).
    """
    if context is None:
        context = ExecutionContext(mode=ExecutionMode.GLOBAL)

    effective_dry_run = request.dry_run or context.dry_run

    processed:         List[ProcessedRecord] = []
    total_corrected    = 0
    total_invalid      = 0
    total_gender_inf   = 0
    total_category_inf = 0

    for idx, record in enumerate(request.records):
        validations: List[ValidationDetail] = []
        inferences:  List[InferenceDetail]  = []

        farmer_name    = record.farmer_name
        contact_number = record.contact_number
        mgnrega_card   = record.mgnrega_card
        gender         = record.gender.strip().lower()
        category       = record.category.strip().lower()

        # 1. Name case
        if context.allow_rule("format_name") and context.allow_column("farmer_name"):
            nr = _name_fmt.format_name(farmer_name)
            if nr.was_changed and not effective_dry_run:
                farmer_name = nr.formatted
            validations.append(ValidationDetail(
                field="farmer_name", original=nr.original, corrected=nr.formatted,
                was_changed=nr.was_changed, is_valid=True, reason=nr.description,
            ))
        else:
            validations.append(ValidationDetail(
                field="farmer_name", original=farmer_name, corrected=farmer_name,
                was_changed=False, is_valid=True,
                reason="Skipped by execution context",
            ))

        # 2. Contact
        if context.allow_rule("validate_contact") and context.allow_column("contact_number"):
            cr = _contact_v.validate(contact_number)
            if cr.is_valid and cr.was_changed and not effective_dry_run:
                contact_number = cr.corrected
            validations.append(ValidationDetail(
                field="contact_number", original=cr.original, corrected=cr.corrected,
                was_changed=cr.was_changed, is_valid=cr.is_valid, reason=cr.reason,
            ))
        else:
            validations.append(ValidationDetail(
                field="contact_number", original=contact_number, corrected=contact_number,
                was_changed=False, is_valid=True,
                reason="Skipped by execution context",
            ))

        # 3. MGNREGA
        if context.allow_rule("validate_mgnrega") and context.allow_column("mgnrega_card"):
            mr = _mgnrega_v.validate(mgnrega_card)
            if mr.is_valid and mr.was_corrected and not effective_dry_run:
                mgnrega_card = mr.card
            validations.append(ValidationDetail(
                field="mgnrega_card", original=mr.original, corrected=mr.card,
                was_changed=mr.was_corrected, is_valid=mr.is_valid, reason=mr.reason,
            ))
        else:
            validations.append(ValidationDetail(
                field="mgnrega_card", original=mgnrega_card, corrected=mgnrega_card,
                was_changed=False, is_valid=True,
                reason="Skipped by execution context",
            ))

        # 4. Gender inference
        if (not gender
                and context.allow_rule("infer_gender")
                and context.allow_column("gender")):
            g = _lib_infer_gender(farmer_name, spouse_or_father=record.father_or_spouse)
            if g.is_determined and not effective_dry_run:
                gender = g.gender
            inferences.append(InferenceDetail(
                field="gender", value=g.gender, confidence=g.confidence,
                tier=g.tier, label=g.label, reason=g.reason,
            ))
            if g.is_determined:
                total_gender_inf += 1

        # 5. Category inference
        if (not category
                and context.allow_rule("infer_category")
                and context.allow_column("category")):
            c = _brlf_infer_category(farmer_name)
            if c.is_determined and not effective_dry_run:
                category = c.value
            inferences.append(InferenceDetail(
                field="category", value=c.value, confidence=c.confidence,
                tier=c.tier, label=c.label, reason=c.reason,
            ))
            if c.is_determined:
                total_category_inf += 1

        has_corrections    = any(v.was_changed    for v in validations)
        has_invalid_fields = any(not v.is_valid   for v in validations)
        has_inferences     = bool(inferences)

        if has_corrections:    total_corrected += 1
        if has_invalid_fields: total_invalid   += 1

        processed.append(ProcessedRecord(
            row_index=idx, farmer_name=farmer_name,
            contact_number=contact_number, mgnrega_card=mgnrega_card,
            gender=gender, category=category,
            validations=validations, inferences=inferences,
            has_corrections=has_corrections,
            has_invalid_fields=has_invalid_fields,
            has_inferences=has_inferences,
        ))

    return ProcessResponse(
        total=len(processed), corrected=total_corrected, invalid=total_invalid,
        inferred_gender=total_gender_inf, inferred_category=total_category_inf,
        records=processed,
    )


# ─────────────────────────────────────────────────────────────────────────────
# FILE-LEVEL PIPELINE  (BulkProcessor + /api/upload jobs)
# ─────────────────────────────────────────────────────────────────────────────

class DataMigrationEngine:
    """
    Full file migration pipeline:
      1.  Backup original
      2.  Load (.xlsx / .xls / .csv / .tsv)
      3.  Detect & quarantine placeholder IDs
      4.  Validate survey dates
      5.  Deduplicate (sort DESC by SubmissionDate → keep most recent)
      6.  Reattach placeholders
      7.  Auto-correct gender / category (vectorised)
      8.  Spell-check name columns
      9.  Validate output
      10. Save cleaned file
      11. Build 9-sheet audit report
    """

    def __init__(self, config: dict):
        self.config          = config
        self.unique_key      = config.get("unique_key_column",      "Unique Key")
        self.farmer_col      = config.get("farmer_name_column",     "farmer_name")
        self.father_col      = config.get("father_spouse_column",   "father_spouse_name")
        self.date_col        = config.get("submission_date_column", "SubmissionDate")
        self.survey_date_col = config.get("survey_date_column",     "date_of_survey")
        self.district_col    = config.get("district_column",        "district")   # v4.1: for Tier-4 geo prior
        self.mgnrega_col     = config.get("mgnrega_column",         "mgnrega")
        self.correct_fields  = config.get("auto_correct_fields",    ["gender", "category"])
        self.keep_dup        = config.get("keep_duplicate",         "first")
        self.min_confidence  = config.get("min_confidence",         75)

        spell_cfg            = config.get("spell_check", {})
        self.spell_enabled   = spell_cfg.get("enabled", True)
        self.spell_cols      = spell_cfg.get("columns", [self.farmer_col, self.father_col])
        self._spell_checker  = NameSpellChecker(
            min_confidence    = spell_cfg.get("min_confidence",    70),
            high_confidence   = spell_cfg.get("high_confidence",   80),
            max_edit_distance = spell_cfg.get("max_edit_distance", 2),
        )

        self._changes:             list = []
        self._date_issues:         list = []
        self._placeholder_records: list = []
        self._spell_corrections:   list = []

    # ── Public API ────────────────────────────────────────────────────────────

    def process(
        self,
        input_path: str,
        output_dir: str,
        context:    Optional[ExecutionContext] = None,
    ) -> dict:
        """
        Run full pipeline. Returns summary dict.
        context controls execution mode; None → GLOBAL (v4.0 backward-compatible).
        """
        if context is None:
            context = ExecutionContext(mode=ExecutionMode.GLOBAL)

        if context.mode == ExecutionMode.ADVISORY:
            context.dry_run = True

        log.info(
            "Execution Mode=%s | dry_run=%s | Columns=%s | Rules=%s",
            context.mode.value,
            context.dry_run,
            sorted(context.selected_columns) if context.selected_columns else "ALL",
            sorted(context.selected_rules)   if context.selected_rules   else "ALL",
        )

        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = Path(input_path).stem
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # 1. Backup
        backup_path = (Path(output_dir) / "backup" /
                       f"{stem}_BACKUP_{ts}{Path(input_path).suffix}")
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(input_path, backup_path)
        log.info("Backup: %s", backup_path)

        # 2. Load
        df_original, sheet_name = self._load(input_path)
        n_original = len(df_original)
        log.info("Loaded %d records from '%s'", n_original, sheet_name)

        # 3. Placeholder detection
        df_main, df_placeholders = self._split_placeholders(df_original)
        n_placeholder = len(df_placeholders)
        if n_placeholder:
            log.warning("%d placeholder-ID records quarantined", n_placeholder)

        # 4. Date validation
        self._validate_survey_dates(df_original)
        log.info("Date issues: %d", len(self._date_issues))

        # 5. Dedup (real records only) — gated by context
        if context.allow_rule("deduplicate"):
            df_deduped, n_removed = self._remove_duplicates(df_main)
            log.info("Duplicates removed: %d → %d records", n_removed, len(df_deduped))
        else:
            df_deduped = df_main.copy()
            n_removed  = 0
            log.info("Dedup skipped by execution context")

        # 6. Reattach placeholders
        df_combined = (pd.concat([df_deduped, df_placeholders], ignore_index=True)
                       if n_placeholder else df_deduped)

        # ── Column integrity snapshot (SELECTIVE mode only) ───────────────────
        if context.mode == ExecutionMode.SELECTIVE:
            frozen_columns  = set(df_combined.columns) - set(context.selected_columns or set())
            frozen_snapshot = df_combined[list(frozen_columns)].copy() if frozen_columns else None
        else:
            frozen_columns  = set()
            frozen_snapshot = None

        # 7. Auto-correct — gated by context
        if context.allow_rule("auto_correct"):
            df_corrected = self._auto_correct(df_combined, context)
            log.info("Auto-corrections: %d", len(self._changes))
        else:
            df_corrected = df_combined.copy()
            log.info("Auto-correct skipped by execution context")

        # 8. Name case formatting — gated by context
        if context.allow_rule("format_name"):
            df_corrected = self._format_name_columns(df_corrected, context)
            log.info("Name case formatting complete")
        else:
            log.info("Name case formatting skipped by execution context")

        # 9. Spell check — gated by context
        if self.spell_enabled and context.allow_rule("spell_check"):
            df_corrected = self._spell_check_names(df_corrected, context)
            log.info("Spell corrections: %d", len(self._spell_corrections))
        else:
            if not self.spell_enabled:
                log.info("Spell check disabled in config")
            else:
                log.info("Spell check skipped by execution context")

        # 10. Output validation
        validation_warnings = self._validate_output(
            df_original, df_corrected, n_removed, n_placeholder
        )
        for w in validation_warnings:
            log.warning("VALIDATION: %s", w)

        # ── Column integrity guard (SELECTIVE mode) ──────────────────────────
        if frozen_snapshot is not None and not frozen_snapshot.empty:
            current_frozen = df_corrected[list(frozen_columns)]
            if not frozen_snapshot.reset_index(drop=True).equals(
                current_frozen.reset_index(drop=True)
            ):
                raise RuntimeError(
                    "Integrity violation: Non-selected columns were modified "
                    "during SELECTIVE mode execution. "
                    f"Frozen columns: {sorted(frozen_columns)}"
                )
            log.info("Column integrity check passed — frozen columns unchanged ✓")

        # BUG FIX (v4.1-a): Reset index to contiguous 0..N-1 before all output
        # steps. After sort_values + drop_duplicates the index may contain
        # non-contiguous original row positions. All downstream .at[] lookups and
        # _build_review_file rely on a clean RangeIndex.
        df_corrected = df_corrected.reset_index(drop=True)
        df_original  = df_original.reset_index(drop=True)

        # 11. Save FILE 1 — Cleaned
        out_cleaned = Path(output_dir) / f"{stem}_CLEANED_{ts}.xlsx"
        df_corrected.to_excel(out_cleaned, index=False)
        log.info(f"File 1 — Cleaned {'(Advisory/Unchanged)' if context.dry_run else '(Modified)'}: %s", out_cleaned)

        # 12. Save FILE 2 — Merged
        out_merged = Path(output_dir) / f"{stem}_MERGED_{ts}.xlsx"
        df_merged = self._build_merged(df_corrected) if not context.dry_run else df_corrected
        df_merged.to_excel(out_merged, index=False)
        log.info("File 2 — Merged (%d rows): %s", len(df_merged), out_merged)

        # 13. Save FILE 3 — Review
        out_review = Path(output_dir) / f"{stem}_REVIEW_{ts}.xlsx"
        self._build_review_file(df_original, df_corrected, out_review)
        log.info("File 3 — Review (highlighted): %s", out_review)

        # 14. Report
        out_report = Path(output_dir) / f"{stem}_REPORT_{ts}.xlsx"
        self._build_report(
            df_original, df_corrected, out_report,
            n_removed, n_placeholder, ts, input_path, validation_warnings,
        )
        log.info("Report: %s", out_report)

        # 15. JSON Suggestions Backup
        import json
        out_suggestions = Path(output_dir) / "suggestions.json"
        with open(out_suggestions, "w", encoding="utf-8") as f:
            json.dump({
                "changes": self._changes,
                "spell_changes": self._spell_corrections
            }, f, indent=2)
        log.info("Suggestions raw data: %s", out_suggestions)

        return {
            "input_file":          input_path,
            "sheet":               sheet_name,
            "original_records":    n_original,
            "placeholder_records": n_placeholder,
            "duplicates_removed":  n_removed,
            "records_after_dedup": len(df_deduped),
            "auto_corrections":    len(self._changes),
            "spell_corrections":   len(self._spell_corrections),
            "date_issues":         len(self._date_issues),
            "final_records":       len(df_corrected),
            "merged_records":      len(df_merged),
            "validation_warnings": validation_warnings,
            "backup_file":         str(backup_path),
            "output_file":         str(out_cleaned),
            "merged_file":         str(out_merged),
            "review_file":         str(out_review),
            "report_file":         str(out_report),
            "changes":             self._changes,
            "spell_changes":       self._spell_corrections,
        }

    # ── Pipeline steps ────────────────────────────────────────────────────────

    def _load(self, path: str):
        ext = Path(path).suffix.lower()
        if ext in (".xlsx", ".xls"):
            xf = pd.ExcelFile(path)
            sn = xf.sheet_names[0]
            return pd.read_excel(path, sheet_name=sn), sn
        if ext == ".csv":
            return pd.read_csv(path), "CSV"
        if ext == ".tsv":
            return pd.read_csv(path, sep="\t"), "TSV"
        raise ValueError(f"Unsupported file type: {ext}")

    def _split_placeholders(self, df: pd.DataFrame):
        if self.unique_key not in df.columns:
            return df.copy(), pd.DataFrame(columns=df.columns)
        is_ph   = df[self.unique_key].apply(_clean_val).isin(PLACEHOLDER_IDS)
        df_ph   = df[is_ph].copy()
        df_real = df[~is_ph].copy()
        for _, row in df_ph.iterrows():
            self._placeholder_records.append({
                "Farmer_Name":    row.get(self.farmer_col, ""),
                "Village":        row.get("Village Name", ""),
                "Bad_ID":         row.get(self.unique_key, ""),
                "SubmissionDate": row.get(self.date_col, ""),
                "Suggested_Key":  self._make_surrogate_key(row),
            })
        return df_real, df_ph

    def _make_surrogate_key(self, row) -> str:
        farmer  = _clean_val(row.get(self.farmer_col, "")).replace(" ", "_")
        village = _clean_val(row.get("Village Name", "")).replace(" ", "_")
        date    = _clean_val(str(row.get(self.date_col, ""))).replace(" ", "_")[:10]
        return f"SURR_{farmer}__{village}__{date}"

    def _validate_survey_dates(self, df: pd.DataFrame) -> None:
        if self.survey_date_col not in df.columns:
            return
        for i, val in df[self.survey_date_col].items():
            if not _clean_val(val):
                continue
            if pd.isna(pd.to_datetime(val, errors="coerce")):
                self._date_issues.append({
                    "Row":           i + 2,
                    "Farmer_Name":   df.at[i, self.farmer_col] if self.farmer_col in df.columns else "",
                    "Village":       df.at[i, "Village Name"]  if "Village Name"  in df.columns else "",
                    "Bad_Value":     val,
                    "Suggested_Fix": f"Use '{self.date_col}' column value as fallback",
                })

    def _remove_duplicates(self, df: pd.DataFrame):
        if self.unique_key not in df.columns:
            log.warning("Column '%s' not found — skipping dedup", self.unique_key)
            return df.copy(), 0
        n_before = len(df)
        df = df.copy()
        if self.date_col in df.columns:
            sort_col = "_sort_date"
            had_col  = sort_col in df.columns
            df[sort_col] = pd.to_datetime(df[self.date_col], errors="coerce")
            df = df.sort_values(sort_col, ascending=False)
            if not had_col:
                df = df.drop(columns=[sort_col])
        else:
            log.warning("Column '%s' not found — dedup order not guaranteed", self.date_col)
        df_out = (df.drop_duplicates(subset=[self.unique_key], keep="first")
                    .reset_index(drop=True))
        return df_out, n_before - len(df_out)

    def _auto_correct(self, df: pd.DataFrame, context: ExecutionContext) -> pd.DataFrame:
        """
        Vectorised gender and category correction.

        BUG FIX (v4.1-a): reset_index(drop=True) at top ensures index is always
        0..N-1 before any label-based .at[] operations. The enumerate(results)
        pattern is also replaced with results.items() so the loop variable is the
        actual label (idx) not a positional counter — these diverge on merged
        multi-file DataFrames where concat() preserves original file indices.
        """
        # BUG FIX (v4.1-a): guarantee clean RangeIndex before any .at[] ops
        df = df.reset_index(drop=True)
        df = df.copy()
        self._changes = []

        VALID_GENDERS    = {"male", "female", "others"}
        VALID_CATEGORIES = {"obc", "st", "sc", "gen", "sbc", "pvtg"}

        NORM_GENDER = {
            "male":"male","m":"male","Male":"male","MALE":"male","M":"male",
            "female":"female","f":"female","Female":"female","FEMALE":"female","F":"female",
            "others":"others","other":"others","Others":"others",
        }
        NORM_CATEGORY = {
            "obc":"obc","OBC":"obc","Obc":"obc",
            "st":"st","ST":"st","St":"st",
            "sc":"sc","SC":"sc","Sc":"sc",
            "gen":"gen","GEN":"gen","general":"gen","General":"gen","GENERAL":"gen",
            "sbc":"sbc","SBC":"sbc",
            "pvtg":"pvtg","PVTG":"pvtg",
        }

        # ── GENDER ────────────────────────────────────────────────────────────
        if (
            context.allow_column("gender")
            and context.allow_rule("auto_correct_gender")
            and "gender" in self.correct_fields
            and "gender" in df.columns
        ):
            def _resolve_gender(row):
                raw  = "" if _clean_val(row.get("gender", "")) == "" else str(row.get("gender", "")).strip()
                norm = NORM_GENDER.get(raw, _clean_val(raw))
                pred, conf, reason = self._infer_gender(
                    row.get(self.farmer_col, ""), row.get(self.father_col, "")
                )
                if pred and conf >= self.min_confidence and pred != norm:
                    return (pred, conf, "Name Analysis", reason)
                if norm in VALID_GENDERS and norm != raw:
                    return (norm, 99, "Normalisation", f"Standardised '{raw}' → '{norm}'")
                return (norm if norm in VALID_GENDERS else _clean_val(raw), None, None, None)

            results = df.apply(_resolve_gender, axis=1)
            # BUG FIX (v4.1-a): results.items() → idx IS the label, not positional counter
            for idx, (new_val, conf, method, reason) in results.items():
                raw_val = df.at[idx, "gender"]
                raw_str = "" if _clean_val(raw_val) == "" else str(raw_val).strip()
                if new_val and new_val != raw_str and conf is not None:
                    snap = df.loc[idx].copy()
                    if not context.dry_run:
                        df.at[idx, "gender"] = new_val
                    self._log_change(idx + 2, snap, "gender",
                                     _clean_val(raw_val), new_val, conf, method, reason)
                elif new_val and new_val != raw_str and not context.dry_run:
                    df.at[idx, "gender"] = new_val

        # ── CATEGORY ──────────────────────────────────────────────────────────
        if (
            context.allow_column("category")
            and context.allow_rule("auto_correct_category")
            and "category" in self.correct_fields
            and "category" in df.columns
        ):
            def _resolve_category(row):
                raw  = str(row.get("category", "") or "").strip()
                norm = NORM_CATEGORY.get(raw, _clean_val(raw))
                if norm in VALID_CATEGORIES:
                    if norm != raw:
                        return (norm, 99, "Normalisation", f"Standardised '{raw}' → '{norm}'")
                    return (norm, None, None, None)
                # v4.1: pass district for Tier-4 geographic prior
                district = str(row.get(self.district_col, "") or "")
                pred, conf, reason = self._infer_category(
                    row.get(self.farmer_col, ""), district
                )
                if pred and conf >= CAT_MIN_CONFIDENCE:
                    return (pred, conf, "Surname Heuristic", reason)
                return (_clean_val(raw), None, None, None)

            results = df.apply(_resolve_category, axis=1)
            # BUG FIX (v4.1-a): results.items() — same fix as gender block
            for idx, (new_val, conf, method, reason) in results.items():
                raw_val = df.at[idx, "category"]
                raw_str = str(raw_val).strip() if raw_val else ""
                if new_val and new_val != raw_str and conf is not None:
                    snap = df.loc[idx].copy()
                    if not context.dry_run:
                        df.at[idx, "category"] = new_val
                    self._log_change(idx + 2, snap, "category",
                                     _clean_val(raw_val), new_val, conf, method, reason)

        # ── Majority vote within Unique Key groups ────────────────────────────
        if self.unique_key in df.columns:
            for fld in self.correct_fields:
                if fld not in df.columns:
                    continue
                if not context.allow_column(fld):
                    log.debug("Majority vote skipped for column '%s' (not in context)", fld)
                    continue
                for key, grp in df.groupby(self.unique_key):
                    if str(key).strip().lower() in PLACEHOLDER_IDS:
                        log.info(
                            "Majority vote skipped for placeholder key '%s' "
                            "(%d records) — not a real duplicate group",
                            key, len(grp),
                        )
                        continue
                    vals = grp[fld].dropna().astype(str).str.strip().str.lower()
                    if vals.nunique() <= 1:
                        continue
                    counts   = vals.value_counts()
                    majority = counts.index[0]
                    total    = len(vals)
                    maj_n    = int(counts.iloc[0])
                    pct      = maj_n / total
                    if pct < 0.60:
                        continue
                    others = {v: int(c) for v, c in counts.iloc[1:].items()}
                    reason = (f"Key '{key}': '{majority}' is {maj_n}/{total} "
                              f"({pct*100:.0f}%). Minority: {others}")
                    conf   = 74 if pct < 0.80 else 82
                    snaps  = {i: df.loc[i].copy() for i in grp.index}
                    for i in grp.index:
                        cur = _clean_val(df.at[i, fld])
                        if cur != majority:
                            if not context.dry_run:
                                df.at[i, fld] = majority
                            self._log_change(i + 2, snaps[i], fld,
                                             cur, majority, conf, "Majority Vote", reason)
        return df

    def _format_name_columns(
        self, df: pd.DataFrame, context: ExecutionContext
    ) -> pd.DataFrame:
        df = df.copy()
        name_cols = [c for c in [self.farmer_col, self.father_col] if c in df.columns]
        total_fixed = 0
        for col in name_cols:
            if not context.allow_column(col):
                log.info("Name formatting skipped for column '%s' (not in context)", col)
                continue
            _, changes = _name_fmt.format_dataframe_column(df, col)
            if changes and not context.dry_run:
                for ch in changes:
                    df.at[ch["DataFrame_Index"], col] = ch["Formatted"]
            total_fixed += len(changes)
        log.info("Name case formatter: %d name(s) corrected across %d column(s)",
                 total_fixed, len(name_cols))
        return df

    def _spell_check_names(
        self, df: pd.DataFrame, context: ExecutionContext
    ) -> pd.DataFrame:
        df = df.copy()
        self._spell_corrections = []
        for col in self.spell_cols:
            if col not in df.columns:
                log.warning("Spell-check column '%s' not found", col)
                continue
            if not context.allow_column(col):
                log.info("Spell check skipped for column '%s' (not in context)", col)
                continue
            results = self._spell_checker.check_series(df[col])
            for i, row in results.iterrows():
                if not row["changed"]:
                    continue
                if not context.dry_run:
                    df.at[i, col] = row["corrected"]
                for tok in row["token_results"]:
                    if (tok["status"] in ("corrected", "review")
                            and tok["original"].lower() != tok["corrected"].lower()):
                        self._spell_corrections.append({
                            "Excel_Row":        i + 2,
                            "Column":           col,
                            "Full_Name_Before": row["original"],
                            "Full_Name_After":  row["corrected"],
                            "Token_Position":   tok["role"],
                            "Original_Token":   tok["original"],
                            "Corrected_Token":  tok["corrected"],
                            "Confidence_%":     tok["confidence"],
                            "Method":           tok["method"],
                            "Status":           tok["status"],
                        })
        return df

    def _validate_output(
        self,
        df_original: pd.DataFrame,
        df_clean:    pd.DataFrame,
        n_duplicates_removed: int,
        n_placeholder: int = 0,
    ) -> list:
        warnings = []
        expected = len(df_original) - n_duplicates_removed
        actual   = len(df_clean)
        if actual != expected:
            warnings.append(
                f"Record count mismatch: expected {expected:,} "
                f"(original {len(df_original):,} − dedup {n_duplicates_removed:,}) "
                f"but got {actual:,}"
            )

        orig_cols  = set(df_original.columns)
        clean_cols = set(df_clean.columns)
        if orig_cols - clean_cols:
            warnings.append(f"Columns dropped: {orig_cols - clean_cols}")
        if clean_cols - orig_cols:
            warnings.append(f"Unexpected new columns: {clean_cols - orig_cols}")

        if "gender" in df_clean.columns:
            VALID_G = {"male", "female", "others", ""}
            bad = df_clean["gender"].apply(_clean_val)
            bv  = [v for v in bad[~bad.isin(VALID_G)].unique() if v not in ("", "nan")]
            if bv:
                warnings.append(f"{len(bv)} invalid gender values remain: {bv}")

        if "category" in df_clean.columns:
            VALID_C = {"obc", "st", "sc", "gen", "sbc", "pvtg", ""}
            bad = df_clean["category"].apply(_clean_val)
            bv  = [v for v in bad[~bad.isin(VALID_C)].unique() if v not in ("", "nan")]
            if bv:
                warnings.append(f"{len(bv)} invalid category values remain: {bv}")

        if not warnings:
            log.info("Output validation passed ✓")
        return warnings

    def _log_change(self, row_num, row, field, old, new, conf, method, reason):
        self._changes.append({
            "Excel_Row":     row_num,
            "Farmer_Name":   row.get(self.farmer_col, ""),
            "Father_Spouse": row.get(self.father_col, ""),
            "Field":         field,
            "Old_Value":     old,
            "New_Value":     new,
            "Confidence_%":  conf,
            "Method":        method,
            "Reason":        reason,
        })

    @staticmethod
    def _infer_gender(farmer_name: str, father_spouse: str = "") -> tuple:
        """Delegates to gender_lib — single source of truth."""
        r = _lib_infer_gender(farmer_name, spouse_or_father=father_spouse)
        return (r.gender, r.confidence, r.reason)

    @staticmethod
    def _infer_category(farmer_name: str, district: str = "") -> tuple:
        """
        Delegates to check_intelligence.inference.engine — single source of truth.
        v4.1: passes district for Tier-4 geographic prior (Gadchiroli, Nandurbar…).
        """
        r = _brlf_infer_category(farmer_name, district=district)
        return (r.value, r.confidence, r.reason)

    # ── File 2: Merged output ─────────────────────────────────────────────────

    def _build_merged(self, df: pd.DataFrame) -> pd.DataFrame:
        uk       = self.unique_key
        date_col = self.date_col

        if uk not in df.columns:
            log.warning("Unique key column '%s' not found — merged = cleaned", uk)
            return df.copy()

        df = df.copy()
        sort_col = "_merge_sort_date"
        if date_col in df.columns:
            df[sort_col] = pd.to_datetime(df[date_col], errors="coerce")
            df = df.sort_values(sort_col, ascending=False)
            df = df.drop(columns=[sort_col])

        merged = df.drop_duplicates(subset=[uk], keep="first").reset_index(drop=True)
        log.info("Merged: %d → %d rows (kept most-recent per '%s')",
                 len(df), len(merged), uk)
        return merged

    # ── File 3: Review output (color-highlighted) ─────────────────────────────

    _REVIEW_COLORS = {
        "gender":             "D6E4FF",
        "category":           "FFF2CC",
        "farmer_name":        "E2EFDA",
        "father_spouse_name": "E2EFDA",
        "contact_number":     "FCE4D6",
        "mgnrega":            "F4CCFF",
        "date_of_survey":     "FFE2E2",
    }
    _NULL_COLOR  = "FFF2CC"
    _DEDUP_COLOR = "F2F2F2"

    def _build_review_file(
        self,
        df_original:  pd.DataFrame,
        df_corrected: pd.DataFrame,
        out_path:     Path,
    ) -> None:
        """
        FILE 3 — Review: df_corrected with every changed cell highlighted.

        Color scheme:
          🔵 Blue     (#D6E4FF) — gender changed
          🟡 Yellow   (#FFF2CC) — category changed
          🟢 Green    (#E2EFDA) — name (farmer / father) changed
          🟠 Orange   (#FCE4D6) — contact number changed
          🟣 Purple   (#F4CCFF) — mgnrega changed
          🔴 Red      (#FFE2E2) — date issue / null value in key field
          ⚫ Grey     (#F2F2F2) — duplicate row (present in cleaned, absent in merged)
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        THIN = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        changed_cells: dict = {}
        n_rows = len(df_corrected)

        # 1. Auto-correct changes (gender, category)
        for ch in self._changes:
            df_idx = ch["Excel_Row"] - 2
            if df_idx < 0 or df_idx >= n_rows:
                continue
            field = ch["Field"]
            color = self._REVIEW_COLORS.get(field, "FFFACD")
            changed_cells[(df_idx, field)] = color

        # 2. Spell corrections
        for sc in self._spell_corrections:
            df_idx = sc["Excel_Row"] - 2
            if df_idx < 0 or df_idx >= n_rows:
                continue
            col   = sc["Column"]
            color = self._REVIEW_COLORS.get(col, "E2EFDA")
            changed_cells[(df_idx, col)] = color

        # 3. Name case changes — positional .iat[] — immune to index issues
        compare_len = min(len(df_original), len(df_corrected))
        for name_col in [self.farmer_col, self.father_col]:
            if name_col not in df_original.columns or name_col not in df_corrected.columns:
                continue
            color    = self._REVIEW_COLORS.get(name_col, "E2EFDA")
            orig_ser = df_original[name_col].fillna("").astype(str).str.strip()
            corr_ser = df_corrected[name_col].fillna("").astype(str).str.strip()
            for pos in range(compare_len):
                if orig_ser.iat[pos] != corr_ser.iat[pos]:
                    changed_cells[(pos, name_col)] = color

        # 4. Contact number changes
        if "contact_number" in df_original.columns and "contact_number" in df_corrected.columns:
            orig_ser = df_original["contact_number"].fillna("").astype(str).str.strip()
            corr_ser = df_corrected["contact_number"].fillna("").astype(str).str.strip()
            for pos in range(compare_len):
                if orig_ser.iat[pos] != corr_ser.iat[pos]:
                    changed_cells[(pos, "contact_number")] = self._REVIEW_COLORS["contact_number"]

        # 5. MGNREGA changes
        if self.mgnrega_col in df_original.columns and self.mgnrega_col in df_corrected.columns:
            orig_ser = df_original[self.mgnrega_col].fillna("").astype(str).str.strip()
            corr_ser = df_corrected[self.mgnrega_col].fillna("").astype(str).str.strip()
            for pos in range(compare_len):
                if orig_ser.iat[pos] != corr_ser.iat[pos]:
                    changed_cells[(pos, self.mgnrega_col)] = self._REVIEW_COLORS.get("mgnrega", "F4CCFF")

        # 6. Date issues
        if self.survey_date_col in df_corrected.columns:
            for issue in self._date_issues:
                df_idx = issue.get("Row", 2) - 2
                if 0 <= df_idx < n_rows:
                    changed_cells[(df_idx, self.survey_date_col)] = self._REVIEW_COLORS["date_of_survey"]

        # 7. Null gender / category cells
        for null_col in ["gender", "category"]:
            if null_col not in df_corrected.columns:
                continue
            ser = df_corrected[null_col]
            for pos in range(len(df_corrected)):
                val = ser.iat[pos]
                if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                    changed_cells[(pos, null_col)] = self._NULL_COLOR

        # 8. Identify rows that survive into merged output
        merged_positions: set = set()
        if self.unique_key in df_corrected.columns:
            tmp = df_corrected[[self.unique_key] + (
                [self.date_col] if self.date_col in df_corrected.columns else []
            )].copy()
            tmp["_pos"] = range(len(tmp))
            if self.date_col in tmp.columns:
                tmp["_sort"] = pd.to_datetime(tmp[self.date_col], errors="coerce")
                tmp = tmp.sort_values("_sort", ascending=False)
            first_per_key  = tmp.drop_duplicates(subset=[self.unique_key], keep="first")["_pos"]
            merged_positions = set(first_per_key)

        # ── Build workbook ────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Review"

        cols    = list(df_corrected.columns)
        col_idx = {c: i + 1 for i, c in enumerate(cols)}

        # Legend row (row 1)
        legend_items = [
            ("Gender changed",    "D6E4FF"),
            ("Category changed",  "FFF2CC"),
            ("Name corrected",    "E2EFDA"),
            ("Contact corrected", "FCE4D6"),
            ("MGNREGA corrected", "F4CCFF"),
            ("Date issue",        "FFE2E2"),
            ("Missing value",     "FFF2CC"),
            ("Duplicate row",     "F2F2F2"),
        ]
        ws.row_dimensions[1].height = 20
        for li, (label, hex_color) in enumerate(legend_items):
            ci   = li * 2 + 1
            cell = ws.cell(1, ci, f"  {label}  ")
            cell.fill      = PatternFill("solid", fgColor=hex_color)
            cell.font      = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = THIN
            if ci + 1 <= len(cols):
                ws.merge_cells(start_row=1, start_column=ci,
                               end_row=1,   end_column=ci + 1)

        # Header row (row 2)
        ws.row_dimensions[2].height = 30
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(2, ci, col_name)
            cell.fill      = PatternFill("solid", fgColor="1F4E79")
            cell.font      = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN

        # Data rows (row 3+)
        df_vals = df_corrected.values
        for pos in range(len(df_corrected)):
            excel_row    = pos + 3
            is_duplicate = (
                self.unique_key in df_corrected.columns and pos not in merged_positions
            )
            for col_name in cols:
                ci    = col_idx[col_name]
                value = df_vals[pos, ci - 1]

                if (pos, col_name) in changed_cells:
                    bg = changed_cells[(pos, col_name)]
                elif is_duplicate:
                    bg = "F2F2F2"
                else:
                    bg = "FFFFFF"

                if isinstance(value, float) and pd.isna(value):
                    value = ""
                elif isinstance(value, float) and value == int(value):
                    if col_name == "contact_number":
                        value = str(int(value))

                cell = ws.cell(excel_row, ci, value)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.font      = Font(size=9, italic=is_duplicate)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.border    = THIN

        # Column widths
        KEY_COLS = {
            "farmer_name": 28, "father_spouse_name": 26, "gender": 10,
            "category": 10, "contact_number": 14, self.mgnrega_col: 24,
            "date_of_survey": 18, "SubmissionDate": 22, self.unique_key: 22,
            "Village Name": 20, "district": 16, "block": 16,
        }
        for ci, col_name in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = KEY_COLS.get(col_name, 14)

        ws.freeze_panes = "A3"
        wb.save(out_path)

        # BUG FIX (v4.1-b): was 'merged_indices' (NameError) — correct name is merged_positions
        log.info(
            "Review file: %d rows, %d changed cells, %d duplicate rows greyed",
            len(df_corrected),
            len(changed_cells),
            len(df_corrected) - len(merged_positions),
        )

    # ── Report ────────────────────────────────────────────────────────────────

    def _build_report(
        self, orig_df, clean_df, out_path,
        n_duplicates_removed, n_placeholder,
        ts, input_path, validation_warnings,
    ):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        changes_df     = pd.DataFrame(self._changes)
        date_issues_df = pd.DataFrame(self._date_issues)
        placeholder_df = pd.DataFrame(self._placeholder_records)
        spell_df       = pd.DataFrame(self._spell_corrections)

        wb = Workbook()
        wb.remove(wb.active)

        DARK   = "1F4E79"; MID    = "2E75B6"; LIGHT  = "BDD7EE"
        GREEN  = "E2EFDA"; YELLOW = "FFF2CC"; ORANGE = "FCE4D6"
        RED    = "FFE2E2"; WHITE  = "FFFFFF"; GRAY   = "F2F2F2"
        THIN   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def H(ws, r, c, v, bg=DARK, fg=WHITE, bold=True, sz=11, span=1):
            cell = ws.cell(r, c, v)
            cell.font      = Font(color=fg, bold=bold, size=sz)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN
            if span > 1:
                ws.merge_cells(
                    f"{get_column_letter(c)}{r}:{get_column_letter(c+span-1)}{r}"
                )
            return cell

        def C(ws, r, c, v="", bg=WHITE, bold=False, align="left", wrap=False, fg="000000"):
            cell = ws.cell(r, c, v)
            cell.font      = Font(bold=bold, size=10, color=fg)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            cell.border    = THIN
            return cell

        def conf_bg(pct):
            try:
                pct = float(pct)
            except (TypeError, ValueError):
                return WHITE
            if pct >= 90: return GREEN
            if pct >= 75: return YELLOW
            if pct >= 60: return ORANGE
            return RED

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws1 = wb.create_sheet("📊 Summary")
        ws1.row_dimensions[1].height = 38
        H(ws1, 1, 1,
          f"DATA MIGRATION REPORT v4.1 — {Path(input_path).name}",
          sz=14, span=3)
        C(ws1, 2, 1,
          f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}  "
          f"|  Engine: DataMigrationEngine v4.1",
          bg=GRAY)
        ws1.merge_cells("A2:C2")

        rows = [
            ("", ""),
            ("── INPUT ──", ""),
            ("Source File",               Path(input_path).name),
            ("Original Records",          f"{len(orig_df):,}"),
            ("Columns",                   len(orig_df.columns)),
            ("", ""),
            ("── DATA QUALITY FLAGS ──", ""),
            ("Placeholder ID Records",    f"{n_placeholder:,}"),
            ("Date Field Issues",         f"{len(self._date_issues):,}"),
            ("", ""),
            ("── CLEANING ──", ""),
            ("Duplicates Removed",        f"{n_duplicates_removed:,}"),
            ("Records After Dedup",       f"{len(orig_df)-n_duplicates_removed:,}"),
            ("Auto-Corrections",          f"{len(self._changes):,}"),
            ("Name Spelling Corrections", f"{len(self._spell_corrections):,}"),
            ("", ""),
            ("── OUTPUT ──", ""),
            ("Final Records",             f"{len(clean_df):,}"),
            ("Reduction %",
             f"{(1-len(clean_df)/max(len(orig_df),1))*100:.1f}%"),
            ("Validation Status",
             "⚠ WARNINGS" if validation_warnings else "✓ PASSED"),
        ]

        if len(changes_df):
            rows += [("", ""), ("── CORRECTIONS BY FIELD ──", "")]
            for fld, cnt in changes_df["Field"].value_counts().items():
                rows.append((f"  {fld}", f"{cnt:,}"))
            rows += [("", ""), ("── CORRECTION METHODS ──", "")]
            for mth, cnt in changes_df["Method"].value_counts().items():
                rows.append((f"  {mth}", f"{cnt:,}"))

        for ri, (label, value) in enumerate(rows, start=4):
            is_sec = label.startswith("──")
            bg = MID if is_sec else WHITE
            fg = WHITE if is_sec else "000000"
            for ci, txt in enumerate([label, value], 1):
                cell = ws1.cell(ri, ci, txt)
                cell.font      = Font(bold=is_sec, color=fg, size=10)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(
                    horizontal="left" if ci == 1 else "center",
                    vertical="center",
                )
                cell.border = THIN

        ws1.column_dimensions["A"].width = 32
        ws1.column_dimensions["B"].width = 22

        if validation_warnings:
            start = 4 + len(rows) + 2
            H(ws1, start, 1, "⚠ VALIDATION WARNINGS", bg="C00000", span=2)
            for ri, w in enumerate(validation_warnings, start + 1):
                C(ws1, ri, 1, w, bg=RED, wrap=True)
                ws1.merge_cells(f"A{ri}:B{ri}")

        # ── Sheets 2–8 ────────────────────────────────────────────────────────
        self._write_changes_sheets(
            wb, changes_df, date_issues_df, placeholder_df,
            H, C, conf_bg, get_column_letter,
            orig_df, clean_df, n_duplicates_removed, n_placeholder, input_path,
            DARK, LIGHT, ORANGE, RED, GREEN, YELLOW, WHITE, GRAY, THIN,
        )

        # ── Sheet 9: Name Corrections ─────────────────────────────────────────
        ws9 = wb.create_sheet("✏ Name Corrections")
        H(ws9, 1, 1,
          "NAME SPELLING CORRECTIONS  (🟢 auto ≥80%  |  🟡 review 70-79%)",
          bg=MID, span=10)

        SCOLS  = ["Excel_Row", "Column", "Full_Name_Before", "Full_Name_After",
                  "Token_Position", "Original_Token", "Corrected_Token",
                  "Confidence_%", "Method", "Status"]
        SWIDTHS = [9, 18, 32, 32, 12, 18, 18, 13, 22, 12]

        if len(spell_df) == 0 or not all(c in spell_df.columns for c in SCOLS):
            ws9.cell(2, 1, "✓  No spelling corrections needed.")
        else:
            for ci, col in enumerate(SCOLS, 1):
                H(ws9, 2, ci, col.replace("_", " "), bg=LIGHT, fg=DARK, sz=10)
            for ri, row in enumerate(spell_df[SCOLS].itertuples(index=False), 3):
                status = row[9]
                bg = GREEN if status == "corrected" else YELLOW
                for ci, v in enumerate(row, 1):
                    cell = ws9.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=bg)
                    cell.border    = THIN
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(wrap_text=(ci in (3, 4, 9)), vertical="top")
            for ci, w in enumerate(SWIDTHS, 1):
                ws9.column_dimensions[get_column_letter(ci)].width = w
            n_auto   = len(spell_df[spell_df["Status"] == "corrected"])
            n_review = len(spell_df[spell_df["Status"] == "review"])
            fr       = len(spell_df) + 4
            C(ws9, fr, 1,
              f"Total: {len(spell_df)} correction(s) across "
              f"{spell_df['Excel_Row'].nunique()} row(s).  "
              f"Auto: {n_auto}  |  Review: {n_review}",
              bg=LIGHT, bold=True, wrap=True)
            ws9.merge_cells(f"A{fr}:{get_column_letter(len(SCOLS))}{fr}")

        wb.save(out_path)

    def _write_changes_sheets(
        self, wb, changes_df, date_issues_df, placeholder_df,
        H, C, conf_bg, gcl,
        orig_df, clean_df, n_removed, n_placeholder, input_path,
        DARK, LIGHT, ORANGE, RED, GREEN, YELLOW, WHITE, GRAY, THIN,
    ):
        from openpyxl.styles import Font, PatternFill, Alignment

        # Sheet 2 — All Changes
        ws2    = wb.create_sheet("🔄 All Changes")
        COLS   = ["Excel_Row","Farmer_Name","Father_Spouse","Field",
                  "Old_Value","New_Value","Confidence_%","Method","Reason"]
        WIDTHS = [9, 28, 26, 10, 12, 12, 14, 24, 62]
        if len(changes_df) == 0:
            ws2.cell(1, 1, "✓  No corrections needed — data was already clean.")
        else:
            H(ws2, 1, 1, "ALL AUTO-CORRECTIONS", span=len(COLS))
            for ci, col in enumerate(COLS, 1):
                H(ws2, 2, ci, col.replace("_", " "), bg=LIGHT, fg=DARK, sz=10)
            for ri, row in enumerate(changes_df[COLS].itertuples(index=False), 3):
                bg = conf_bg(row[6])
                for ci, v in enumerate(row, 1):
                    cell = ws2.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=bg)
                    cell.alignment = Alignment(wrap_text=(ci == 9), vertical="top")
                    cell.border    = THIN
                    cell.font      = Font(size=9)
            for ci, w in enumerate(WIDTHS, 1):
                ws2.column_dimensions[gcl(ci)].width = w

        # Sheet 3 — Gender
        ws3 = wb.create_sheet("⚤ Gender Changes")
        gc  = (changes_df[changes_df["Field"] == "gender"]
               if len(changes_df) else pd.DataFrame())
        if len(gc) == 0:
            ws3.cell(1, 1, "✓  No gender corrections needed.")
        else:
            GCOLS = ["Excel_Row","Farmer_Name","Father_Spouse",
                     "Old_Value","New_Value","Confidence_%","Method","Reason"]
            for ci, col in enumerate(GCOLS, 1):
                H(ws3, 1, ci, col.replace("_", " "), bg=LIGHT, fg=DARK)
            for ri, row in enumerate(gc[GCOLS].itertuples(index=False), 2):
                bg = conf_bg(row[5])
                for ci, v in enumerate(row, 1):
                    cell = ws3.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=bg)
                    cell.border    = THIN
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(wrap_text=(ci == 8), vertical="top")
            for ci, w in enumerate([9, 28, 26, 12, 12, 14, 24, 62], 1):
                ws3.column_dimensions[gcl(ci)].width = w

        # Sheet 4 — Category
        ws4 = wb.create_sheet("🏷 Category Changes")
        cc  = (changes_df[changes_df["Field"] == "category"]
               if len(changes_df) else pd.DataFrame())
        if len(cc) == 0:
            ws4.cell(1, 1, "✓  No category corrections needed.")
        else:
            CCOLS = ["Excel_Row","Farmer_Name","Old_Value","New_Value",
                     "Confidence_%","Method","Reason"]
            for ci, col in enumerate(CCOLS, 1):
                H(ws4, 1, ci, col.replace("_", " "), bg=LIGHT, fg=DARK)
            for ri, row in enumerate(cc[CCOLS].itertuples(index=False), 2):
                bg = conf_bg(row[4])
                for ci, v in enumerate(row, 1):
                    cell = ws4.cell(ri, ci, v)
                    cell.fill   = PatternFill("solid", fgColor=bg)
                    cell.border = THIN
                    cell.font   = Font(size=9)
            for ci, w in enumerate([9, 28, 12, 12, 14, 24, 62], 1):
                ws4.column_dimensions[gcl(ci)].width = w

        # Sheet 5 — Review Needed
        ws5 = wb.create_sheet("⚠ Review Needed")
        low = (changes_df[changes_df["Confidence_%"] < 75]
               if len(changes_df) else pd.DataFrame())
        H(ws5, 1, 1,
          "Changes below 75% confidence — verify before finalising",
          bg=ORANGE, fg="843C0C", span=7)
        if len(low) == 0:
            ws5.cell(2, 1, "✓  All changes high-confidence (≥75%).")
        else:
            RCOLS = ["Excel_Row","Farmer_Name","Field","Old_Value",
                     "New_Value","Confidence_%","Reason"]
            for ci, col in enumerate(RCOLS, 1):
                H(ws5, 2, ci, col.replace("_", " "), bg=ORANGE, fg="843C0C")
            for ri, row in enumerate(low[RCOLS].itertuples(index=False), 3):
                for ci, v in enumerate(row, 1):
                    cell = ws5.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=ORANGE)
                    cell.border    = THIN
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(wrap_text=(ci == 7), vertical="top")
            for ci, w in enumerate([9, 28, 10, 12, 12, 14, 62], 1):
                ws5.column_dimensions[gcl(ci)].width = w

        # Sheet 6 — Duplicate Summary
        ws6 = wb.create_sheet("🗂 Duplicate Summary")
        H(ws6, 1, 1, "DUPLICATE RECORDS REMOVED", span=3)
        for ri, (label, value, bg) in enumerate([
            ("Total records in original:", f"{len(orig_df):,}",           WHITE),
            ("Placeholder ID records:",    f"{n_placeholder:,}",          YELLOW),
            ("Unique records kept:",       f"{len(orig_df)-n_removed:,}", GREEN),
            ("Duplicate records removed:", f"{n_removed:,}",              YELLOW),
            ("Duplication rate:",
             f"{n_removed/max(len(orig_df),1)*100:.1f}%",                 WHITE),
        ], start=2):
            C(ws6, ri, 1, label, bold=True)
            C(ws6, ri, 2, value, bg=bg, align="center")
        C(ws6, 8, 1,
          f"Strategy: Sort by '{self.date_col}' DESC → keep first occurrence "
          f"of each '{self.unique_key}' (most recent kept)",
          bg=GRAY, wrap=True)
        ws6.merge_cells("A8:C8")
        ws6.column_dimensions["A"].width = 34
        ws6.column_dimensions["B"].width = 18

        # Sheet 7 — Date Issues
        ws7 = wb.create_sheet("📅 Date Issues")
        H(ws7, 1, 1,
          f"'{self.survey_date_col}' contains non-date values",
          bg="C00000", span=5)
        if len(date_issues_df) == 0:
            ws7.cell(2, 1, "✓  All date values are valid.")
        else:
            DCOLS = ["Row","Farmer_Name","Village","Bad_Value","Suggested_Fix"]
            for ci, col in enumerate(DCOLS, 1):
                H(ws7, 2, ci, col.replace("_", " "), bg=LIGHT, fg=DARK)
            for ri, row in enumerate(date_issues_df[DCOLS].itertuples(index=False), 3):
                for ci, v in enumerate(row, 1):
                    cell = ws7.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=RED)
                    cell.border    = THIN
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(wrap_text=(ci == 5), vertical="top")
            for ci, w in enumerate([8, 28, 22, 32, 48], 1):
                ws7.column_dimensions[gcl(ci)].width = w

        # Sheet 8 — Placeholder IDs
        ws8 = wb.create_sheet("🆔 Placeholder IDs")
        H(ws8, 1, 1,
          f"Records with placeholder '{self.unique_key}' values",
          bg=ORANGE, fg="843C0C", span=5)
        if len(placeholder_df) == 0:
            ws8.cell(2, 1, "✓  No placeholder IDs found.")
        else:
            PCOLS = ["Farmer_Name","Village","Bad_ID","SubmissionDate","Suggested_Key"]
            for ci, col in enumerate(PCOLS, 1):
                H(ws8, 2, ci, col.replace("_", " "), bg=ORANGE, fg="843C0C")
            for ri, row in enumerate(placeholder_df[PCOLS].itertuples(index=False), 3):
                for ci, v in enumerate(row, 1):
                    cell = ws8.cell(ri, ci, v)
                    cell.fill      = PatternFill("solid", fgColor=YELLOW)
                    cell.border    = THIN
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(wrap_text=(ci == 5), vertical="top")
            for ci, w in enumerate([28, 22, 14, 22, 48], 1):
                ws8.column_dimensions[gcl(ci)].width = w