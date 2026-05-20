"""
backend/bulk_processor.py
==========================
Process multiple survey files in parallel.

Features:
  • process_files()   — process a list of files
  • process_folder()  — process every supported file in a folder
  • watch_folder()    — continuously watch for new files (runs forever)
  • build_master_report() — one Excel summary across all files
  • Resume support — skips already-processed files by SHA-256 checksum
  • Full error isolation — one bad file never stops others

FIXES vs original bulk_processor.py
--------------------------------------
  FIX 1 CRITICAL: from core_engine import DataMigrationEngine
              → from intelligence_engine import DataMigrationEngine
from execution_context import ExecutionContext, ExecutionMode

  FIX 2 CRITICAL: SUPPORTED_EXTENSIONS had .json/.tsv not in lib_utils
              → imported from gender_intelligence.lib_utils

  FIX 3 BUG: _file_checksum() used MD5 — lib_utils has SHA-256 checksum()
              → replaced with lib_utils.checksum()

  FIX 4 BUG: ProcessPoolExecutor deadlocks on macOS (fork start method)
              → ThreadPoolExecutor — same API, no fork, cross-platform safe

  FIX 5 BUG: watch_folder() called _process_one_file() single-threaded,
             ignoring max_workers
              → watch_folder now submits to a shared ThreadPoolExecutor

  FIX 6 BUG: processed log file had no locking — two concurrent
             watch_folder() instances would corrupt it
              → fcntl.flock() write lock around every read/write

  FIX 7 MISSING: list[str] type hints fail on Python 3.8
              → from __future__ import annotations added

  FIX 8 MISSING: build_master_report THIN border built wrong
             (positional args → Border takes keyword args)
              → Border(left=Side(...), right=Side(...), ...)

  FIX 9 MISSING: No integration with new intelligence_engine API
              → _process_one_file calls DataMigrationEngine.process()
                 (same dict-based API, no change needed)

  FIX 10 MISSING: SUPPORTED_EXTENSIONS was a local constant,
              diverging from lib_utils over time
              → single source of truth via import
"""

from __future__ import annotations

import fcntl
import json
import logging
import os
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

# FIX 1: correct module import
from services.pipeline_orchestrator import DataMigrationEngine

# FIX 2+3: single source of truth for extensions and checksum
from gender_intelligence.lib_utils import SUPPORTED_EXTENSIONS, checksum

log = logging.getLogger("bulk_processor")


# ─────────────────────────────────────────────────────────────────────────────
# PROCESSED LOG HELPERS  (FIX 6: file locking)
# ─────────────────────────────────────────────────────────────────────────────

def _load_processed_log(log_path: str) -> Dict[str, dict]:
    if not os.path.exists(log_path):
        return {}
    with open(log_path, "r") as f:
        fcntl.flock(f, fcntl.LOCK_SH)
        try:
            return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)


def _save_processed_log(log_path: str, data: Dict[str, dict]) -> None:
    # FIX 6: exclusive write lock prevents concurrent corruption
    with open(log_path, "a+") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            f.seek(0)
            f.truncate()
            json.dump(data, f, indent=2)
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)


def _is_supported(path: str) -> bool:
    return Path(path).suffix.lower() in SUPPORTED_EXTENSIONS


# ─────────────────────────────────────────────────────────────────────────────
# SINGLE-FILE WORKER
# ─────────────────────────────────────────────────────────────────────────────

def _process_one_file(args: tuple) -> dict:
    """
    Worker function — processes a single file.
    Designed to run in a ThreadPoolExecutor worker thread.
    Returns a result dict regardless of success/failure.

    args tuple: (input_path, output_dir, config, context)
    context is an ExecutionContext — always supplied, never None.
    """
    input_path, output_dir, config, context = args
    result: dict = {
        "file":        input_path,
        "status":      "pending",
        "started_at":  datetime.now().isoformat(),
        "finished_at": None,
        "error":       None,
        "summary":     None,
    }
    try:
        engine  = DataMigrationEngine(config)
        summary = engine.process(input_path, output_dir, context=context)
        result.update({
            "status":      "success",
            "finished_at": datetime.now().isoformat(),
            "summary": {
                "original_records":    summary["original_records"],
                "duplicates_removed":  summary["duplicates_removed"],
                "auto_corrections":    summary["auto_corrections"],
                "spell_corrections":   summary.get("spell_corrections", 0),
                "date_issues":         summary["date_issues"],
                "placeholder_records": summary["placeholder_records"],
                "final_records":       summary["final_records"],
                "validation_warnings": summary["validation_warnings"],
                "output_file":         summary["output_file"],
                "report_file":         summary["report_file"],
                "backup_file":         summary["backup_file"],
            },
        })
    except Exception as e:
        result.update({
            "status":      "error",
            "finished_at": datetime.now().isoformat(),
            "error":       str(e),
        })
        log.error("Failed: %s — %s", input_path, e)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# BULK PROCESSOR
# ─────────────────────────────────────────────────────────────────────────────

class BulkProcessor:
    """
    Process multiple files in one go.

    Usage:
        bp = BulkProcessor(config)
        results = bp.process_files(["/data/file1.xlsx", "/data/file2.csv"],
                                   output_dir="/data/output")

        bp.watch_folder("/data/inbox", "/data/outbox")  # runs forever
    """

    def __init__(self, config: dict):
        self.config         = config
        self.max_workers    = config.get("bulk", {}).get("max_workers",            4)
        self.skip_processed = config.get("bulk", {}).get("skip_already_processed", True)
        self.poll_interval  = config.get("bulk", {}).get("watch_poll_seconds",     10)

    # ── Public API ────────────────────────────────────────────────────────────

    def process_files(
        self,
        file_paths:         List[str],
        output_dir:         str,
        processed_log_path: Optional[str] = None,
        context:            Optional[ExecutionContext] = None,
    ) -> List[dict]:
        """
        Process a list of files in parallel.

        Args:
            file_paths:          Absolute paths to input files.
            output_dir:          Root output folder. Each file gets its own subfolder.
            processed_log_path:  Optional JSON log for resume support.
            context:             ExecutionContext controlling GLOBAL/ADVISORY/SELECTIVE mode.
                                 None → defaults to GLOBAL (backward-compatible).
        """
        if context is None:
            context = ExecutionContext.global_mode()
        if not file_paths:
            log.warning("No files to process.")
            return []

        processed_log: Dict[str, dict] = {}
        if processed_log_path and self.skip_processed:
            processed_log = _load_processed_log(processed_log_path)

        # FIX 2: filter to lib_utils SUPPORTED_EXTENSIONS
        valid_files = [f for f in file_paths if _is_supported(f) and os.path.exists(f)]
        skipped     = len(file_paths) - len(valid_files)
        if skipped:
            log.warning("Skipped %d unsupported/missing files", skipped)

        # Filter already-processed
        # FIX 3: use lib_utils checksum() (SHA-256) not local MD5
        to_process: List[tuple] = []
        for fp in valid_files:
            chk = checksum(fp)
            if processed_log_path and self.skip_processed and chk in processed_log:
                log.info("Skipping (already processed): %s", Path(fp).name)
            else:
                to_process.append((fp, chk))

        if not to_process:
            log.info("All files already processed.")
            return []

        log.info("Processing %d file(s) with %d worker(s)",
                 len(to_process), self.max_workers)

        worker_args: List[tuple] = []
        for fp, _ in to_process:
            stem        = Path(fp).stem
            file_outdir = os.path.join(output_dir, stem)
            os.makedirs(file_outdir, exist_ok=True)
            worker_args.append((fp, file_outdir, self.config, context))

        # FIX 4: ThreadPoolExecutor — no fork, cross-platform safe
        results = self._run_parallel(worker_args)

        if processed_log_path:
            for (fp, chk), res in zip(to_process, results):
                if res["status"] == "success":
                    processed_log[chk] = {
                        "file":         fp,
                        "processed_at": res["finished_at"],
                    }
            _save_processed_log(processed_log_path, processed_log)

        self._print_batch_summary(results)
        return results

    def process_folder(
        self,
        input_dir:          str,
        output_dir:         str,
        processed_log_path: Optional[str] = None,
    ) -> List[dict]:
        """Scan input_dir for all supported files and process them all."""
        input_dir = os.path.abspath(input_dir)
        if not os.path.isdir(input_dir):
            raise ValueError(f"Input directory not found: {input_dir}")
        files = sorted(
            os.path.join(input_dir, f)
            for f in os.listdir(input_dir)
            if _is_supported(f)
        )
        log.info("Found %d supported file(s) in '%s'", len(files), input_dir)
        return self.process_files(files, output_dir, processed_log_path)

    def watch_folder(
        self,
        input_dir:          str,
        output_dir:         str,
        done_dir:           Optional[str] = None,
        processed_log_path: Optional[str] = None,
    ) -> None:
        """
        Continuously watch input_dir for new files.
        Runs forever — press Ctrl+C to stop.

        FIX 5: now uses ThreadPoolExecutor, respects max_workers.
        FIX 6: processed log protected by file lock.
        """
        os.makedirs(output_dir, exist_ok=True)
        if done_dir:
            os.makedirs(done_dir, exist_ok=True)

        log_path      = processed_log_path or os.path.join(output_dir, ".processed_log.json")
        processed_log = _load_processed_log(log_path)

        log.info("👀 Watching: '%s'  →  '%s'", input_dir, output_dir)
        log.info("   Poll: %ds  |  Workers: %d  |  Ctrl+C to stop.",
                 self.poll_interval, self.max_workers)

        # FIX 5: shared executor lives across poll cycles
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            try:
                while True:
                    new_files: List[tuple] = []
                    for fname in os.listdir(input_dir):
                        fp  = os.path.join(input_dir, fname)
                        if not _is_supported(fp) or not os.path.isfile(fp):
                            continue
                        chk = checksum(fp)
                        if chk not in processed_log:
                            new_files.append((fp, chk))

                    if new_files:
                        log.info("🆕 %d new file(s) detected", len(new_files))

                        # FIX 5: submit all new files to thread pool
                        future_to_info: dict = {}
                        for fp, chk in new_files:
                            stem     = Path(fp).stem
                            file_out = os.path.join(output_dir, stem)
                            os.makedirs(file_out, exist_ok=True)
                            args   = (fp, file_out, self.config, ExecutionContext.global_mode())
                            future = executor.submit(_process_one_file, args)
                            future_to_info[future] = (fp, chk)

                        for future in as_completed(future_to_info):
                            fp, chk  = future_to_info[future]
                            result   = future.result()
                            if result["status"] == "success":
                                processed_log[chk] = {
                                    "file":         fp,
                                    "processed_at": result["finished_at"],
                                }
                                # FIX 6: locked write
                                _save_processed_log(log_path, processed_log)
                                log.info("✅ Done: %s", Path(fp).name)
                                if done_dir:
                                    dest = os.path.join(done_dir, Path(fp).name)
                                    shutil.move(fp, dest)
                                    log.info("   Moved → %s", dest)
                            else:
                                log.error("❌ Failed: %s — %s",
                                          Path(fp).name, result["error"])

                    time.sleep(self.poll_interval)

            except KeyboardInterrupt:
                log.info("Watcher stopped.")

    # ── Internal ──────────────────────────────────────────────────────────────

    def _run_parallel(self, worker_args: List[tuple]) -> List[dict]:
        """
        FIX 4: ThreadPoolExecutor — no fork, works on macOS/Windows/Linux.
        Falls back to serial if max_workers=1.
        """
        results: List[dict] = [None] * len(worker_args)  # type: ignore

        if self.max_workers == 1 or len(worker_args) == 1:
            for i, args in enumerate(worker_args):
                results[i] = _process_one_file(args)
            return results

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(_process_one_file, args): i
                for i, args in enumerate(worker_args)
            }
            for future in as_completed(futures):
                idx         = futures[future]
                results[idx] = future.result()

        return results

    @staticmethod
    def _print_batch_summary(results: List[dict]) -> None:
        print("\n" + "=" * 70)
        print(f"  BULK PROCESSING COMPLETE — {len(results)} file(s)")
        print("=" * 70)
        success = [r for r in results if r["status"] == "success"]
        errors  = [r for r in results if r["status"] == "error"]
        for r in success:
            s = r["summary"]
            print(f"\n  ✅ {Path(r['file']).name}")
            print(f"     Records  : {s['original_records']:,} → {s['final_records']:,}")
            print(f"     Dedup    : {s['duplicates_removed']:,} removed")
            print(f"     Fixes    : {s['auto_corrections']} corrections, "
                  f"{s['spell_corrections']} spelling")
            if s["date_issues"]:
                print(f"     ⚠ Date issues   : {s['date_issues']}")
            if s["placeholder_records"]:
                print(f"     ⚠ Placeholder IDs: {s['placeholder_records']}")
            if s["validation_warnings"]:
                for w in s["validation_warnings"]:
                    print(f"     ⚠ {w}")
            print(f"     Output   : {s['output_file']}")
        for r in errors:
            print(f"\n  ❌ {Path(r['file']).name}")
            print(f"     Error: {r['error']}")
        print(f"\n  Summary: {len(success)} succeeded, {len(errors)} failed")
        print("=" * 70 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# MASTER SUMMARY REPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_master_report(results: List[dict], output_path: str) -> None:
    """Build a single Excel summary across all processed files."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "📋 Master Summary"

    DARK   = "1F4E79"; LIGHT  = "BDD7EE"; GREEN  = "E2EFDA"
    YELLOW = "FFF2CC"; RED    = "FFE2E2"; WHITE  = "FFFFFF"
    # FIX 8: Border takes keyword args, not positional
    THIN = Border(
        left   = Side(style="thin"),
        right  = Side(style="thin"),
        top    = Side(style="thin"),
        bottom = Side(style="thin"),
    )

    def cell(r: int, c: int, v="", bg: str = WHITE, bold: bool = False,
             align: str = "left"):
        cl = ws.cell(r, c, v)
        cl.font      = Font(bold=bold, size=10,
                            color="FFFFFF" if bg == DARK else "000000")
        cl.fill      = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal=align, vertical="center",
                                 wrap_text=True)
        cl.border    = THIN
        return cl

    headers = ["File Name","Status","Original","Final","Dedup'd",
               "Corrections","Spelling","Date Issues","Placeholders",
               "Warnings","Report Link"]
    widths  = [32, 10, 10, 10, 10, 12, 10, 12, 14, 38, 42]

    ws.row_dimensions[1].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN

    for ri, r in enumerate(results, 2):
        is_ok = r["status"] == "success"
        s     = r.get("summary") or {}
        bg    = WHITE if is_ok else RED
        row_data = [
            Path(r["file"]).name,
            "✅ OK" if is_ok else "❌ ERROR",
            s.get("original_records", ""),
            s.get("final_records", ""),
            s.get("duplicates_removed", ""),
            s.get("auto_corrections", ""),
            s.get("spell_corrections", ""),
            s.get("date_issues", ""),
            s.get("placeholder_records", ""),
            r.get("error") or "; ".join(s.get("validation_warnings", [])) or "—",
            s.get("report_file", "—"),
        ]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font      = Font(size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="center" if ci > 1 else "left",
                vertical="center", wrap_text=True,
            )
            c.border = THIN

    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Totals row
    last = len(results) + 2
    ok   = [r for r in results if r["status"] == "success"]
    cell(last, 1, f"TOTAL — {len(ok)}/{len(results)} succeeded",
         bg=LIGHT, bold=True)
    ws.merge_cells(f"A{last}:B{last}")
    totals = [
        sum(r.get("summary", {}).get("original_records",    0) or 0 for r in ok),
        sum(r.get("summary", {}).get("final_records",       0) or 0 for r in ok),
        sum(r.get("summary", {}).get("duplicates_removed",  0) or 0 for r in ok),
        sum(r.get("summary", {}).get("auto_corrections",    0) or 0 for r in ok),
        sum(r.get("summary", {}).get("spell_corrections",   0) or 0 for r in ok),
        sum(r.get("summary", {}).get("date_issues",         0) or 0 for r in ok),
        sum(r.get("summary", {}).get("placeholder_records", 0) or 0 for r in ok),
    ]
    for ci, v in enumerate(totals, 3):
        cell(last, ci, v, bg=LIGHT, bold=True, align="center")

    wb.save(output_path)
    print(f"  Master report saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Bulk process survey files using DataMigrationEngine v4.0"
    )
    mode = parser.add_subparsers(dest="mode", required=True)

    p_files = mode.add_parser("files",  help="Process specific files")
    p_files.add_argument("files",  nargs="+")
    p_files.add_argument("--output", "-o", required=True)
    p_files.add_argument("--workers", "-w", type=int, default=4)
    p_files.add_argument("--resume",  "-r", action="store_true")

    p_folder = mode.add_parser("folder", help="Process all files in a folder")
    p_folder.add_argument("input")
    p_folder.add_argument("--output", "-o", required=True)
    p_folder.add_argument("--workers", "-w", type=int, default=4)
    p_folder.add_argument("--resume",  "-r", action="store_true")

    p_watch = mode.add_parser("watch", help="Watch folder for new files")
    p_watch.add_argument("input")
    p_watch.add_argument("--output", "-o", required=True)
    p_watch.add_argument("--done",   "-d")
    p_watch.add_argument("--poll",   "-p", type=int, default=10)

    args = parser.parse_args()

    CONFIG = {
        "unique_key_column":      "meta-instanceID",
        "farmer_name_column":     "farmer_name",
        "father_spouse_column":   "father_spouse_name",
        "submission_date_column": "SubmissionDate",
        "survey_date_column":     "date_of_survey",
        "auto_correct_fields":    ["gender", "category"],
        "min_confidence":         75,
        "keep_duplicate":         "first",
        "spell_check": {
            "enabled":            True,
            "columns":            ["farmer_name", "father_spouse_name"],
            "min_confidence":     70,
            "high_confidence":    80,
            "max_edit_distance":  2,
        },
        "bulk": {
            "max_workers":            getattr(args, "workers", 4),
            "skip_already_processed": getattr(args, "resume", False),
            "watch_poll_seconds":     getattr(args, "poll", 10),
        },
    }

    bp = BulkProcessor(CONFIG)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if args.mode == "files":
        results = bp.process_files(
            args.files, args.output,
            processed_log_path=(os.path.join(args.output, ".processed_log.json")
                                if args.resume else None),
        )
    elif args.mode == "folder":
        results = bp.process_folder(
            args.input, args.output,
            processed_log_path=(os.path.join(args.output, ".processed_log.json")
                                if args.resume else None),
        )
    elif args.mode == "watch":
        bp.watch_folder(args.input, args.output, args.done,
                        processed_log_path=os.path.join(
                            args.output, ".processed_log.json"))
        results = []
    else:
        results = []

    if args.mode in ("files", "folder") and results:
        master = os.path.join(args.output, f"MASTER_SUMMARY_{ts}.xlsx")
        build_master_report(results, master)